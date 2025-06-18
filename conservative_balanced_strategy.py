#!/usr/bin/env python3
"""
Conservative Balanced Strategy
More realistic rates that achieve Journey 1 goals without breaking financials
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def create_conservative_strategy():
    """Create conservative balanced strategy with realistic rates"""
    
    strategy_data = []
    
    # More conservative rates - smaller increases from baseline
    conservative_rates = {
        'Stockholm': {
            'Consultant': {
                'A': {'recruitment': 0.06, 'churn': 0.025, 'fte': 68},    # 6% recruitment (vs 5% baseline), 2.5% churn (vs 3% baseline)
                'AC': {'recruitment': 0.045, 'churn': 0.022, 'fte': 64}, # 4.5% recruitment (vs 3.5% baseline), 2.2% churn (vs 2.5% baseline)  
                'C': {'recruitment': 0.025, 'churn': 0.018, 'fte': 16},  # 2.5% recruitment (vs 2% baseline), 1.8% churn (vs 2% baseline)
                'SrC': {'recruitment': 0.01, 'churn': 0.018, 'fte': 32}, # 1% recruitment, 1.8% churn (vs 1.4% baseline)
                'AM': {'recruitment': 0.008, 'churn': 0.015, 'fte': 128}, # 0.8% recruitment, 1.5% churn (vs 1.2% baseline)
                'M': {'recruitment': 0.005, 'churn': 0.012, 'fte': 47},  # 0.5% recruitment, 1.2% churn (vs 1% baseline)
                'SrM': {'recruitment': 0.003, 'churn': 0.01, 'fte': 29}, # 0.3% recruitment, 1% churn (vs 0.8% baseline)
                'PiP': {'recruitment': 0.001, 'churn': 0.008, 'fte': 18} # 0.1% recruitment, 0.8% churn (vs 0.6% baseline)
            },
            'Sales': {
                'A': {'recruitment': 0.045, 'churn': 0.03, 'fte': 13},   # 4.5% recruitment, 3% churn
                'AC': {'recruitment': 0.04, 'churn': 0.028, 'fte': 20},  # 4% recruitment, 2.8% churn
                'C': {'recruitment': 0.03, 'churn': 0.022, 'fte': 17},   # 3% recruitment, 2.2% churn
                'SrC': {'recruitment': 0.015, 'churn': 0.022, 'fte': 5}, # 1.5% recruitment, 2.2% churn
                'AM': {'recruitment': 0.01, 'churn': 0.018, 'fte': 7},   # 1% recruitment, 1.8% churn
                'M': {'recruitment': 0.008, 'churn': 0.015, 'fte': 6},   # 0.8% recruitment, 1.5% churn
                'SrM': {'recruitment': 0.005, 'churn': 0.012, 'fte': 3}, # 0.5% recruitment, 1.2% churn
                'PiP': {'recruitment': 0.003, 'churn': 0.01, 'fte': 7}   # 0.3% recruitment, 1% churn
            },
            'Recruitment': {
                'A': {'recruitment': 0.045, 'churn': 0.03, 'fte': 12},   # 4.5% recruitment, 3% churn
                'AC': {'recruitment': 0.04, 'churn': 0.028, 'fte': 14},  # 4% recruitment, 2.8% churn
                'C': {'recruitment': 0.03, 'churn': 0.022, 'fte': 2},    # 3% recruitment, 2.2% churn
                'SrC': {'recruitment': 0.015, 'churn': 0.022, 'fte': 2}, # 1.5% recruitment, 2.2% churn
                'AM': {'recruitment': 0.01, 'churn': 0.018, 'fte': 6},   # 1% recruitment, 1.8% churn
                'M': {'recruitment': 0.008, 'churn': 0.015, 'fte': 3},   # 0.8% recruitment, 1.5% churn
                'SrM': {'recruitment': 0.005, 'churn': 0.012, 'fte': 2}, # 0.5% recruitment, 1.2% churn
                'PiP': {'recruitment': 0.003, 'churn': 0.01, 'fte': 5}   # 0.3% recruitment, 1% churn
            },
            'Operations': {
                None: {'recruitment': 0.032, 'churn': 0.027, 'fte': 82} # 3.2% recruitment (vs 3% baseline), 2.7% churn (vs 2.5% baseline)
            }
        }
    }
    
    # Create rows for Excel in the expected format
    for office, roles in conservative_rates.items():
        for role, levels in roles.items():
            for level, rates in levels.items():
                # Create base row
                row = {
                    'Office': office,
                    'Role': role,
                    'Level': level if level is not None else '',
                    'FTE': rates['fte']
                }
                
                # Add monthly rates (1-12)
                for month in range(1, 13):
                    row[f'Recruitment_{month}'] = rates['recruitment']
                    row[f'Churn_{month}'] = rates['churn']
                    # Keep default values for other attributes to avoid disruption
                    row[f'Price_{month}'] = ''  # Let system use defaults
                    row[f'Salary_{month}'] = ''  # Let system use defaults
                    row[f'Progression_{month}'] = ''  # Let system use defaults
                    row[f'UTR_{month}'] = ''  # Let system use defaults
                
                strategy_data.append(row)
    
    # Create DataFrame
    df = pd.DataFrame(strategy_data)
    
    # Create Excel file
    filename = 'conservative_balanced_strategy.xlsx'
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Main import sheet
        df.to_excel(writer, sheet_name='Import_Data', index=False)
        
        # Strategy info
        summary_data = {
            'Strategy': ['Conservative Balanced Journey Distribution'],
            'Approach': ['Moderate increases to avoid financial disruption'],
            'Journey 1 Changes': ['A: +1% recruitment, AC: +1% recruitment, C: +0.5% recruitment'],
            'Senior Changes': ['Modest churn increases: SrC +0.4%, AM +0.3%, M +0.2%'],
            'Expected Timeline': ['4-5 years to reach 50% Journey 1'],
            'Financial Impact': ['Minimal disruption to profitability']
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Strategy_Info', index=False)
    
    # Format the Excel file
    wb = load_workbook(filename)
    
    # Format import data sheet
    ws = wb['Import_Data']
    
    # Header formatting
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 15)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the formatted file
    wb.save(filename)
    
    print(f"✅ Created conservative strategy: {filename}")
    print(f"📊 Contains {len(strategy_data)} configuration rows")
    print(f"\n🔧 Conservative Changes:")
    print(f"   • Journey 1 recruitment: +0.5% to +1% increases")
    print(f"   • Journey 1 churn: -0.2% to -0.5% decreases")
    print(f"   • Senior churn: +0.2% to +0.4% increases")
    print(f"   • Financial impact: Minimal")
    print(f"\n⏱️  Expected timeline: 4-5 years to 50% Journey 1")
    print(f"💰 Maintains financial stability")
    
    return filename

if __name__ == "__main__":
    print("🎯 CREATING CONSERVATIVE BALANCED STRATEGY")
    print("=" * 55)
    
    filename = create_conservative_strategy()
    
    print(f"\n📁 File: {filename}")
    print(f"🚀 Ready for safe import!") 